from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


def process_xlsx(file_path: str) -> dict:
    """Extract text from an Excel XLSX file across all sheets."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_text = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_rows = []

            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None]
                if cells:
                    sheet_rows.append(" | ".join(cells))

            if sheet_rows:
                all_text.append(f"--- Sheet: {sheet_name} ---")
                all_text.extend(sheet_rows)

        wb.close()

        text = "\n".join(all_text).strip()
        if not text:
            raise ValueError("No data could be extracted from the XLSX file")

        logger.info(
            f"Extracted data from {len(wb.sheetnames)} sheet(s) in XLSX: {file_path}"
        )
        return {"text": text, "type": "xlsx"}

    except Exception as e:
        logger.error(f"Failed to process XLSX {file_path}: {e}")
        raise
